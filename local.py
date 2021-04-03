# -*- coding =utf-8 -*-
# @Time : 2021/4/3 15:17
# @Author : 胡伊斐
# @File : main.py
# @Software : PyCharm
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string


# 输入参数 str 需要判断的字符串
# 返回值   True：该字符串为浮点数；False：该字符串不是浮点数。
def IsFloatNum(str):
    s = str.split('.')
    if len(s) > 2:
        return False
    else:
        for si in s:
            if not si.isdigit():
                return False
        return True


def findCol(value, sheet):  # 寻找value是第几列,找不到则返回-1
    j = 0
    for col in sheet["2"]:
        if col.value == value:
            return get_column_letter(j + 1)
        j += 1
    return -1

#导入数据
wb = load_workbook("grade.xlsx")
sheet = wb["Sheet1"]

# 定义变量
i = findCol("学分", sheet)#学分所在列
j = findCol("绩点", sheet)#绩点所在列
time = findCol("开课学期", sheet)#开课时间所在列
start = ''
end = ''
grade = []
gradePoint = []
sumgrade = 0
averageGradePoint = []
finalGradePoint = 0

# 将学分信息存入数组
for row in sheet[i]:
    if str(row.value).isdigit() | IsFloatNum(str(row.value)):
        grade.append(row.value)

# 将绩点信息存入数组
for row in sheet[j]:
    if str(row.value).isdigit() | IsFloatNum(str(row.value)):
        gradePoint.append(row.value)

# 将时间信息存入数组
for row in sheet[time]:
    if str(row.value)[-1].isdigit() & (start == ''):
        start = row.value + "学期"
        # print(start)
    elif str(row.value)[-1] != str(start):
        end = str(row.value) + "学期"
        # print(end)

sumgrade = sum(grade)  # 总学分

#计算每门课的加权绩点
for index in range(0, len(grade)):
    s = gradePoint[index] * grade[index] / sumgrade
    averageGradePoint.append(s)

#加权绩点求和得到平均绩点
finalGradePoint = sum(averageGradePoint)

# 输出数据，保存文件（将数据保存到根目录下的gradePoint.xlsx文件里，第一行即为绩点信息）
output = "你在" + str(start) + "到" + str(end) + str(len(grade)) + "门学科的平均绩点为:"
print(output, finalGradePoint)
print("你的绩点数据已经保存到根目录下的gradePoint.xlsx文件里，第一行即为绩点信息")
sheet.insert_rows(1)
sheet["A1"] = output + str(finalGradePoint)
wb.save("gradePoint.xlsx")
input()

